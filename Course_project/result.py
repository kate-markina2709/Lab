import cv2
from skimage.morphology import skeletonize
from PIL import Image
from scipy.spatial import distance
import numpy as np
import math
from math import sqrt
import matplotlib.pyplot as plt
import os
import openpyxl
import os.path

def Clas_element(ii, index_element, massiv_clas, n, k):
    prov = False
    cl = 1
    start = 0
    while prov != True:
        if cl <= k:
            if index_element >= start and index_element < start + n:
                massiv_clas[ii] = cl
                prov = True
            else:
                start = start + n
                cl = cl + 1
        else:
            massiv_clas[ii] = cl
            prov = True
    return massiv_clas

def Check_clas(v1, v2, n, k):
    cl = 1
    start = 0
    while cl <= k:
        if v1 >= start and v1 < start + n and v2 == cl:
            return True
        start = start + n
        cl = cl + 1
    if v1 >= n * k and v2 == k+1:
        return True
    else:
        return False

def histogram(argument):
    im_inp = cv2.imread(argument)
    gray = cv2.cvtColor(im_inp, cv2.COLOR_BGR2GRAY)
    th, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    #print("otsu threshold = ", th)
    bw = bw // 255
    sk = skeletonize(bw)
    out = sk * 255
    cv2.imwrite('result.jpg', out)
    im = Image.open('result.jpg')
    img = np.asarray(im)
    result = np.where(img == 255)
    coords = set()
    for i in range(result[0].size):
        coords.add((result[0][i], result[1][i]))
    #print(*list(coords))
    ii = 0
    k = 0
    corner = [-1000] * (im.size[0] * im.size[1])
    while k < result[0].size - 1:
        # разбиение на квадраты 3*3
        i = result[0][k]
        s = result[1][k] - 1
        while s != im.size[0] - 2:  # ось Х
            if img[i][s + 1] >= 250:
                masx = [0] * 9
                masy = [0] * 9
                masv = [0] * 9
                masx[0] = s + 1
                masx[1] = s
                masx[2] = s + 2
                masx[3] = s + 1
                masx[4] = s
                masx[5] = s + 2
                masx[6] = s + 1
                masx[7] = s
                masx[8] = s + 2
                masy[0] = i
                masy[1] = i
                masy[2] = i
                masy[3] = i - 1
                masy[4] = i - 1
                masy[5] = i - 1
                masy[6] = i + 1
                masy[7] = i + 1
                masy[8] = i + 1
                masv[0] = img[i][s + 1]
                masv[1] = img[i][s]
                masv[2] = img[i][s + 2]
                masv[3] = img[i - 1][s + 1]
                masv[4] = img[i - 1][s]
                masv[5] = img[i - 1][s + 2]
                masv[6] = img[i + 1][s + 1]
                masv[7] = img[i + 1][s]
                masv[8] = img[i + 1][s + 2]
                s = s + 1
                x = np.array([masx[0]])
                y = np.array([masy[0]])
                jj = 1
                kol = 0
                while jj != 9:
                    if masv[jj] >= 250:
                        x = np.append(x, masx[jj])
                        y = np.append(y, masy[jj])
                        kol = kol+1
                    jj = jj + 1
                mx = x.sum()/kol
                my = y.sum()/kol
                a2 = np.dot(x.T, x)/kol
                a11 = np.dot(x.T, y)/kol
                if a2 - mx**22 != 0:
                    kk = (a11 - mx*my)/(a2 - mx**2)
                else:
                    kk = 0
                bb = my - kk*mx
                #ff = np.array([kk*z + bb for z in range(x.shape[0])])
                #plt.plot(ff, c='red')
                #plt.show()
                trox = x[0]
                troy = kk*x[0] + bb
                rxox = x[x.shape[0] - 1]
                rxoy = kk*x[x.shape[0] - 1] + bb
                corn = 0
                if rxox - trox != 0:
                    corn = math.atan2(troy - rxoy, trox - rxox)
                cor = math.degrees(corn)
                corner[ii] = cor
                ii = ii + 1
            else:
                s = s + 1
        k = k + 1
    ind = 0
    while corner[ind] != -1000:
        ind = ind + 1
    corr = [0] * ind
    for i in range(ind):
        corr[i] = corner[i]
    n, bin, patches = plt.hist(corr, bins=18)
    #plt.show()
    return n

def Cosine_distance(a, b):
    if len(a) != len(b):
        return False
    numerator = 0
    denoma = 0
    denomb = 0
    for i in range(len(a)):
        numerator += a[i]*b[i]
        denoma += abs(a[i])**2
        denomb += abs(b[i])**2
    result = 1 - numerator / (sqrt(denoma)*sqrt(denomb))
    return result

def Bhattacharyya_dist(h1, h2):
  def normalize(hh):
    return hh / np.sum(hh)
  return 1 - np.sum(np.sqrt(np.multiply(normalize(h1), normalize(h2))))

def Euclid_dist(h1, h2):
    dst = distance.euclidean(h1, h2)
    return dst

def Resize_image(input_image_path,
                 output_image_path,
                 size):
    original_image = Image.open(input_image_path)
    resized_image = original_image.resize(size)
    resized_image.save(output_image_path)

def Save_hist(pr_name, place):
    name = place + "\p" + str(1) + ".jpg"
    check = place + '.npy'
    check_file = os.path.exists(check)
    if check_file == False:
        Resize_image(name, name, size=(300, 150))
        h1 = histogram(name)
        h = np.array([h1])
        u = 1
        while u < len(pr_name):
            name = place + "\p" + str(u + 1) + ".jpg"
            Resize_image(name, name, size=(300, 150))
            hv = histogram(name)
            h = np.insert(h, [u], [hv], axis=0)
            u = u + 1
        np.save(place, h)

#АЛГОРИТМ, ОСНОВААННЫЙ НА ПОИСКЕ МИНИМАЛЬНОГО РАССТОЯНИЯ МЕЖДУ ЭЛ-МИ
def Find_dist(adres, k_el, n_el):

    def sort_res(masind, n, k):
        clas = [0] * (len(masind))
        real = [0] * (len(masind))
        for ind in range(len(masind)):
            clas = Clas_element(ind, ind, clas, n, k)
        for ind in range(len(masind)):
            real[ind] = Check_clas(masind[ind], clas[ind], n, k)
        s = sum(real)/len(masind)
        return s

    def To_file(i, res, res1, res2, masind, masind1, masind2, result1, result2, result3):
        my_wb = openpyxl.Workbook()
        my_sheet = my_wb.active
        c1 = my_sheet.cell(row=1, column=i)
        c1.value = "Name"
        c2 = my_sheet.cell(row=1, column=i + 1)
        c2.value = "Distance Bh"
        c20 = my_sheet.cell(row=1, column=i + 8)
        c20.value = "Distance Eu"
        c21 = my_sheet.cell(row=1, column=i + 15)
        c21.value = "Distance Cos"
        c3 = my_sheet.cell(row=1, column=i + 2)
        c3.value = "Index"
        c30 = my_sheet.cell(row=1, column=i + 9)
        c30.value = "Index"
        c31 = my_sheet.cell(row=1, column=i + 16)
        c31.value = "Index"
        c5 = my_sheet.cell(row=1, column=i + 3)
        c5.value = "Result"
        c6 = my_sheet.cell(row=2, column=i + 3)
        c6.value = result1
        c50 = my_sheet.cell(row=1, column=i + 10)
        c50.value = "Result"
        c60 = my_sheet.cell(row=2, column=i + 10)
        c60.value = result2
        c51 = my_sheet.cell(row=1, column=i + 17)
        c51.value = "Result"
        c61 = my_sheet.cell(row=2, column=i + 17)
        c61.value = result3
        for ind in range(len(res)):
            c9 = my_sheet.cell(row=ind + 2, column=i)
            c9.value = "p" + str(ind + 1) + ".jpg"
            c7 = my_sheet.cell(row=ind + 2, column=i + 1)
            c7.value = res[ind]
            c70 = my_sheet.cell(row=ind + 2, column=i + 8)
            c70.value = res1[ind]
            c71 = my_sheet.cell(row=ind + 2, column=i + 15)
            c71.value = res2[ind]
            c8 = my_sheet.cell(row=ind + 2, column=i + 2)
            c8.value = masind[ind]
            c80 = my_sheet.cell(row=ind + 2, column=i + 9)
            c80.value = masind1[ind]
            c81 = my_sheet.cell(row=ind + 2, column=i + 16)
            c81.value = masind2[ind]
        my_wb.save("D:\TestMinDist.xlsx")

    pr = os.listdir(adres)
    res_bh = [0] * len(pr)
    masindd = [0] * len(pr)
    res_eu = [0] * len(pr)
    masind_eu = [0] * len(pr)
    res_cos = [0] * len(pr)
    masind_cos = [0] * len(pr)
    Save_hist(pr, adres)
    pl = adres + '.npy'
    h = np.load(pl)
    for tmps in range(len(pr)):
        minel_bh = 1000
        minlen_eu = 100000
        minlen_cos = 1000
        ind1 = 0
        ind2 = 0
        ind3 = 0
        r6 = 0
        r7 = 0
        r5 = 0
        for tmp in range(len(pr)):
            if tmp != tmps:
                r6 = Bhattacharyya_dist(h[tmps], h[tmp])
                if r6 < minel_bh:
                    minel_bh = r6
                    ind1 = tmp
                r7 = Euclid_dist(h[tmps], h[tmp])
                if r7 < minlen_eu:
                    minlen_eu = r7
                    ind2 = tmp
                r5 = Cosine_distance(h[tmps], h[tmp])
                if r5 < minlen_cos:
                    minlen_cos = r5
                    ind3 = tmp
        res_bh[tmps] = r6
        masindd[tmps] = ind1
        res_eu[tmps] = r7
        masind_eu[tmps] = ind2
        res_cos[tmps] = r5
        masind_cos[tmps] = ind3
    result1 = sort_res(masindd, k_el, n_el)
    result2 = sort_res(masind_eu, k_el, n_el)
    result3 = sort_res(masind_cos, k_el, n_el)
    To_file(1, res_bh, res_eu, res_cos, masindd, masind_eu, masind_cos, result1, result2, result3)

#САМО ОБУЧЕНИЕ В РАМКАХ АЛГОРИТМА, ОСНОВАННОГО НА ОБУЧЕНИИ
def Learning(adres, k_el, n_el):

    def Probability(t1, masindex, dist, k, n):
        massiv = [0]*len(dist)
        masivcl = [0]*len(dist)
        count = 0
        s = 0
        for i in range(len(dist)): #правильные классы эл-тов
            massiv = Clas_element(i, i, massiv, k, n)
        for i in range(len(dist)):
            if dist[i] < t1:
                masivcl = Clas_element(i, masindex[i], masivcl, k, n)
                count = count + 1
            else:
                masivcl[i] = n + 1
        if count == len(dist):
            return 0.5
        for i in range(len(dist)):
            if massiv[i] == masivcl[i]:
                s = s + 1
        summa = s/len(dist)
        return summa

    def To_file(dist, index, res_maxv, res_step, res_prob):
        my_wb = openpyxl.Workbook()
        my_sheet = my_wb.active
        c1 = my_sheet.cell(row=1, column=1)
        c1.value = "Name"
        c2 = my_sheet.cell(row=1, column=2)
        c2.value = "Distance"
        c3 = my_sheet.cell(row=1, column=3)
        c3.value = "Index"
        c6 = my_sheet.cell(row=1, column=7)
        c6.value = "Max Value"
        c7 = my_sheet.cell(row=1, column=8)
        c7.value = "Treshold"
        c8 = my_sheet.cell(row=1, column=9)
        c8.value = "Probability"
        for ind in range(len(dist)):
            c9 = my_sheet.cell(row=ind + 2, column=1)
            c9.value = "p" + str(ind + 1) + ".jpg"
            c10 = my_sheet.cell(row=ind + 2, column=2)
            c10.value = dist[ind]
            c11 = my_sheet.cell(row=ind + 2, column=3)
            c11.value = index[ind]
        for ind in range(len(res_maxv)):
            c14 = my_sheet.cell(row=ind + 2, column=7)
            c14.value = res_maxv[ind]
            c15 = my_sheet.cell(row=ind + 2, column=8)
            c15.value = res_step[ind]
            c16 = my_sheet.cell(row=ind + 2, column=9)
            c16.value = res_prob[ind]
        my_wb.save("D:\Result_Teach1.xlsx")

    def C(n, k): #число сочетаний
        if 0 <= k <= n:
            nn = 1
            kk = 1
            for t in range(1, min(k, n - k) + 1):
                nn *= n
                kk *= t
                n -= 1
            return nn // kk
        else:
            return 0

    pr_l = os.listdir(adres)
    Save_hist(pr_l, adres)
    pl = adres + '.npy'
    h_l = np.load(pl)
    val = 0
    distance = [0] * (len(pr_l))
    index_element = [0] * (len(pr_l))
    mat_dist = np.ones([len(pr_l), len(pr_l)], "f")
    x = 0
    while x < len(pr_l):
        cnk = C(k_el, 2)
        vr_dist = [0] * (k_el * k_el + cnk)
        st = [0] * (k_el * k_el + cnk)
        fin = [0] * (k_el * k_el + cnk)
        cnt = 0
        ii = x
        for tmp in range(k_el):
            i = ii + 1
            while i != x + k_el: #для реальных и подделок (они сравниваются м-ду собой)
                st[cnt] = ii
                fin[cnt] = i
                vr_dist[cnt] = Bhattacharyya_dist(h_l[ii], h_l[i])
                cnt = cnt + 1
                i = i + 1
            ii = ii + 1
        ii = x
        if x < n_el*k_el:
            for tmp in range(k_el):  # для подделок
                i = x + n_el * k_el
                while i != x + n_el * k_el + k_el:
                    st[cnt] = ii
                    fin[cnt] = i
                    vr_dist[cnt] = Bhattacharyya_dist(h_l[ii], h_l[i])
                    cnt = cnt + 1
                    i = i + 1
                ii = ii + 1
        minim = 1000
        ind1 = 0
        for tmp in range(len(vr_dist)):
            if vr_dist[tmp] < minim and vr_dist[tmp] != 0:
                minim = vr_dist[tmp]
                ind1 = tmp
        for tmp in range(k_el):
            distance[val] = minim
            index_element[val] = fin[ind1]
            val = val + 1
        x = x + k_el
    for tmps in range(len(pr_l)):
        mat_dist[tmps][tmps] = 0
        for tmp in range(len(pr_l)):
            r6 = Bhattacharyya_dist(h_l[tmps], h_l[tmp])
            mat_dist[tmps][tmp] = r6
    t = 0
    T = (max(max(mat_dist[i]) for i in range(len(pr_l))))
    pro1 = Probability(t, index_element, distance, k_el, n_el)
    res_matrix = np.array([[T, t, pro1]])
    k = 1
    t = t + T/(100*len(pr_l))
    chv = 0
    while t <= T:
        pro = Probability(t, index_element, distance, k_el, n_el)
        res_matrix = np.insert(res_matrix, k, [T, t, pro], axis=0)
        k = k + 1
        if t + T/len(pr_l) > T and chv == 0:
            t = T
            chv = 1
        else:
            t = t + T /(100*len(pr_l))
    res_mat = res_matrix.T
    a_save = adres + "matrix"
    np.save(a_save, res_mat)
    To_file(distance, index_element, res_mat[0], res_mat[1], res_mat[2])
    print('ok')

#ТЕСТИРОВАНИЕ В РАМКАХ АЛГОРИТМА, ОСНОВАННОГО НА ОБУЧЕНИИ
def Testing(adres, k_el, n_el):

    def To_file(tr, res, masr):
        my_wb = openpyxl.Workbook()
        my_sheet = my_wb.active
        c1 = my_sheet.cell(row=1, column=1)
        c1.value = "Treshold"
        c3 = my_sheet.cell(row=2, column=1)
        c3.value = tr
        c2 = my_sheet.cell(row=1, column=2)
        c2.value = "Probability"
        c4 = my_sheet.cell(row=2, column=2)
        c4.value = res
        c5 = my_sheet.cell(row=1, column=5)
        c5.value = "Name"
        c0 = my_sheet.cell(row=1, column=6)
        c0.value = "Class"
        for lon in range(len(masr)):
            c6 = my_sheet.cell(row=lon + 2, column=5)
            c6.value = "p" + str(lon + 1) + ".jpg"
            c7 = my_sheet.cell(row=lon + 2, column=6)
            c7.value = masr[lon]
        my_wb.save("D:\TestTest.xlsx")

    check_file = os.path.exists('D:\Base\Teachmatrix.npy')
    if check_file == False:
        Learning(adres, k_el, n_el)
    data = np.load('D:\Base\Teachmatrix.npy')
    pr_t = os.listdir(adres)
    Save_hist(pr_t, adres)
    znac_treshold = data[1]
    znac_probability = data[2]
    res = 0
    pl = adres + '.npy'
    h_t = np.load(pl)
    vr = np.argmax(znac_probability)
    v_treshold = znac_treshold[vr]
    res_mas = [0] * len(pr_t)
    clas_mas = [0] * len(pr_t)
    for i in range(len(pr_t)):
        clas_mas = Clas_element(i, i, clas_mas, k_el, n_el)
    control = [False] * len(pr_t)
    for tmps in range(len(pr_t)):
        ind1 = 0
        minlen_bh = 1000
        for tmp in range(len(pr_t)):
            if tmp != tmps:
                r6 = Bhattacharyya_dist(h_t[tmps], h_t[tmp])
                if r6 < minlen_bh:
                    minlen_bh = r6
                    ind1 = tmp
        if minlen_bh <= v_treshold:
            res_mas = Clas_element(tmps, ind1, res_mas, k_el, n_el)
        else:
            res_mas[tmps] = n_el + 1
        control[tmps] = Check_clas(tmps, clas_mas[tmps], k_el, n_el)
    for i in range(len(control)):
        if clas_mas[i] == res_mas[i]:
            res = res + 1
    res = res / len(control)
    To_file(v_treshold, res, res_mas)
    print('ok')

#СТАРТ
change_str = input("What do you want to do?\n1. Check authentic images\n2.Training on data\n3. Test on data")
adr = input("Write adres: ")
kol_one_str = input("Enter the number of elements in the class: ")
clas_kol_str = input("Enter the number of classes: ")
kol_one = int(kol_one_str)
clas_kol = int(clas_kol_str)
change = int(change_str)
if change == 1:
    Find_dist(adr, kol_one, clas_kol)
if change == 2:
    Learning(adr, kol_one, clas_kol)
if change == 3:
    Testing(adr, kol_one, clas_kol)
print("That's all")
